# -*- coding: utf-8 -*-
"""
工艺表格自动优化工具

用法：
  python optimize_process_table.py 输入文件.xlsx
  python optimize_process_table.py 输入文件.xlsx 输出文件.xlsx

不传文件时会弹出文件选择窗口。
"""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import column_index_from_string
from openpyxl.utils.cell import coordinate_from_string


HIGHLIGHT_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")
CONDITIONAL_FILL = PatternFill(fill_type="solid", fgColor="E2F0D9")
DEFAULT_TRIGGER_KEYWORDS = ["需复合"]


def app_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def norm_text(value) -> str:
    return "" if value is None else str(value).strip()


def norm_number(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except ValueError:
        return None


def same_number(left, right, tolerance: float = 1e-9) -> bool:
    a = norm_number(left)
    b = norm_number(right)
    if a is None or b is None:
        return False
    return math.isclose(a, b, rel_tol=tolerance, abs_tol=tolerance)


def simplify_next_material(material_name: str) -> str:
    text = norm_text(material_name)
    text = re.sub(r"[（(]\s*复合\s*[）)]", "", text)
    text = text.rstrip("*")
    return text.strip()


def load_trigger_keywords(config_path: Path | None):
    if config_path is None:
        config_path = app_base_dir() / "optimize_rules.json"

    if not config_path.exists():
        config_path.write_text(
            json.dumps(
                {
                    "trigger_keywords": DEFAULT_TRIGGER_KEYWORDS,
                    "note": "trigger_keywords 为空列表 [] 时，任意非空材料名称都可作为第一行。",
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        return DEFAULT_TRIGGER_KEYWORDS

    data = json.loads(config_path.read_text(encoding="utf-8-sig"))
    keywords = data.get("trigger_keywords", DEFAULT_TRIGGER_KEYWORDS)
    if keywords is None:
        return []
    if not isinstance(keywords, list):
        raise ValueError("optimize_rules.json 里的 trigger_keywords 必须是列表。")
    return [str(keyword).strip() for keyword in keywords if str(keyword).strip()]


def base_target_material(material_name: str, trigger_keywords):
    text = norm_text(material_name)
    if not text:
        return None
    if trigger_keywords and not any(keyword in text for keyword in trigger_keywords):
        return None
    if "需复合(" in text:
        return text[: text.find("需复合(") + len("需复合")]
    return text


def find_headers(ws):
    """Return (header_row, material_col, color_col, unit_col, usage_col), or None."""
    max_scan_row = min(ws.max_row, 30)
    for row in range(1, max_scan_row + 1):
        material_col = None
        color_col = None
        unit_col = None
        usage_col = None
        for col in range(1, ws.max_column + 1):
            value = norm_text(ws.cell(row, col).value)
            if value == "材料名称":
                material_col = col
            elif value == "颜色":
                color_col = col
            elif value == "单位":
                unit_col = col
            elif value == "单件用料":
                usage_col = col
        if material_col and color_col and unit_col and usage_col:
            return row, material_col, color_col, unit_col, usage_col
    return None


def find_table_headers(ws):
    """Return header locations needed by optimization and summary."""
    max_scan_row = min(ws.max_row, 30)
    required = {
        "材料名称": None,
        "颜色": None,
        "单位": None,
        "门幅/规格": None,
        "单件用料": None,
        "订单用量": None,
    }
    for row in range(1, max_scan_row + 1):
        found = dict(required)
        for col in range(1, ws.max_column + 1):
            value = norm_text(ws.cell(row, col).value)
            if value in found:
                found[value] = col
        if all(found.values()):
            return {
                "header_row": row,
                "material_col": found["材料名称"],
                "color_col": found["颜色"],
                "unit_col": found["单位"],
                "spec_col": found["门幅/规格"],
                "single_usage_col": found["单件用料"],
                "order_usage_col": found["订单用量"],
            }
    return None


def add_conditional_format(ws, header_row: int, material_col: int):
    first_row = header_row + 1
    last_row = ws.max_row
    if last_row < first_row:
        return
    col_letter = ws.cell(header_row, material_col).column_letter
    data_range = f"{col_letter}{first_row}:{col_letter}{last_row}"
    formula = f'=ISNUMBER(SEARCH("需复合(",{col_letter}{first_row}))'
    rule = FormulaRule(formula=[formula], fill=CONDITIONAL_FILL)
    ws.conditional_formatting.add(data_range, rule)


def unique_output_path(path: Path) -> Path:
    if not path.exists():
        return path
    parent = path.parent
    stem = path.stem
    suffix = path.suffix
    index = 1
    while True:
        candidate = parent / f"{stem}_{index}{suffix}"
        if not candidate.exists():
            return candidate
        index += 1


def merged_top_left_cell(ws, row: int, col: int):
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        return cell
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return ws.cell(merged_range.min_row, merged_range.min_col)
    return cell


def cell_value(ws, row: int, col: int):
    return merged_top_left_cell(ws, row, col).value


def cell_to_number(ws, coordinate: str, seen=None):
    seen = seen or set()
    col_letters, row = coordinate_from_string(coordinate.replace("$", ""))
    col = column_index_from_string(col_letters)
    key = (ws.title, row, col)
    if key in seen:
        return None
    seen.add(key)
    value = cell_value(ws, row, col)
    if isinstance(value, str) and value.startswith("="):
        return evaluate_formula(ws, value, seen)
    return norm_number(value)


def evaluate_formula(ws, formula: str, seen=None):
    expression = formula.strip()
    if expression.startswith("="):
        expression = expression[1:]

    if ":" in expression:
        return None

    def replace_ref(match):
        ref = match.group(0)
        number = cell_to_number(ws, ref, seen)
        if number is None:
            number = 0
        return str(number)

    expression = re.sub(r"\$?[A-Z]{1,3}\$?\d+", replace_ref, expression)
    if not re.fullmatch(r"[0-9eE+\-*/().\s]+", expression):
        return None
    try:
        return float(eval(expression, {"__builtins__": {}}, {}))
    except Exception:
        return None


def displayed_number(ws, row: int, col: int):
    value = cell_value(ws, row, col)
    if isinstance(value, str) and value.startswith("="):
        return evaluate_formula(ws, value)
    return norm_number(value)


def add_order_usage_summary(wb, include_source_rows: bool = True):
    return build_order_usage_summary(
        wb,
        include_source_rows=include_source_rows,
        matched_groups_only=False,
        trigger_keywords=[],
    )


def collect_matched_group_rows(ws, headers, trigger_keywords):
    matched_rows = set()
    row = headers["header_row"] + 1
    while row < ws.max_row:
        material = base_target_material(cell_value(ws, row, headers["material_col"]), trigger_keywords)
        if not material:
            row += 1
            continue

        unit = norm_text(cell_value(ws, row, headers["unit_col"]))
        usage = cell_value(ws, row, headers["single_usage_col"])
        group_rows = [row]
        next_row = row + 1
        while next_row <= ws.max_row:
            next_material = norm_text(cell_value(ws, next_row, headers["material_col"]))
            next_unit = norm_text(cell_value(ws, next_row, headers["unit_col"]))
            next_usage = cell_value(ws, next_row, headers["single_usage_col"])
            if not next_material or next_unit != unit or not same_number(usage, next_usage):
                break
            group_rows.append(next_row)
            next_row += 1

        if len(group_rows) > 1:
            matched_rows.update(group_rows)
            row = next_row
        else:
            row += 1
    return matched_rows


def build_order_usage_summary(
    wb,
    include_source_rows: bool = True,
    matched_groups_only: bool = False,
    trigger_keywords=None,
):
    trigger_keywords = trigger_keywords or []
    summary_name = "订单用量汇总"
    if summary_name in wb.sheetnames:
        del wb[summary_name]

    groups = {}
    for ws in wb.worksheets:
        if ws.title == summary_name:
            continue
        headers = find_table_headers(ws)
        if not headers:
            continue
        matched_rows = (
            collect_matched_group_rows(ws, headers, trigger_keywords)
            if matched_groups_only
            else None
        )

        for row in range(headers["header_row"] + 1, ws.max_row + 1):
            if matched_groups_only and row not in matched_rows:
                continue

            material = norm_text(cell_value(ws, row, headers["material_col"]))
            color = norm_text(cell_value(ws, row, headers["color_col"]))
            unit = norm_text(cell_value(ws, row, headers["unit_col"]))
            spec = norm_text(cell_value(ws, row, headers["spec_col"]))
            order_usage = displayed_number(ws, row, headers["order_usage_col"])

            if not material or not color or not unit or order_usage is None or math.isclose(order_usage, 0):
                continue

            key = (material, color, unit, spec)
            if key not in groups:
                groups[key] = {"total": 0.0, "sources": []}
            groups[key]["total"] += order_usage
            groups[key]["sources"].append(f"{ws.title}!{row}")

    ws = wb.create_sheet(summary_name)
    ws.sheet_view.showGridLines = False

    headers = ["材料名称", "颜色", "单位", "门幅/规格", "订单用量合计"]
    if include_source_rows:
        headers.append("来源行号")

    ws.append(headers)
    for material, color, unit, spec in sorted(groups.keys(), key=lambda item: (item[0], item[1], item[2], item[3])):
        row_data = [material, color, unit, spec, groups[(material, color, unit, spec)]["total"]]
        if include_source_rows:
            row_data.append(", ".join(groups[(material, color, unit, spec)]["sources"]))
        ws.append(row_data)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row in range(2, ws.max_row + 1):
        ws.cell(row, 5).number_format = "0.000000"

    widths = [46, 12, 10, 16, 16, 32]
    for index, width in enumerate(widths[: ws.max_column], 1):
        ws.column_dimensions[ws.cell(1, index).column_letter].width = width
    ws.freeze_panes = "A2"
    return len(groups)


def copy_sheet(source_ws, target_wb, new_title):
    new_ws = target_wb.create_sheet(title=new_title)
    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = new_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.border = cell.border.copy()
                new_cell.alignment = cell.alignment.copy()
                new_cell.number_format = cell.number_format
    for col in source_ws.column_dimensions:
        new_ws.column_dimensions[col].width = source_ws.column_dimensions[col].width
    for row in source_ws.row_dimensions:
        new_ws.row_dimensions[row].height = source_ws.row_dimensions[row].height
    return new_ws


def optimize_sheet(ws, highlight=True, trigger_keywords=None):
    headers = find_headers(ws)
    if not headers:
        return []

    trigger_keywords = trigger_keywords or []
    header_row, material_col, color_col, unit_col, usage_col = headers
    add_conditional_format(ws, header_row, material_col)

    changes = []
    row = header_row + 1
    while row < ws.max_row:
        material_cell = ws.cell(row, material_col)
        usage_cell = ws.cell(row, usage_col)

        material = base_target_material(material_cell.value, trigger_keywords)
        if not material:
            row += 1
            continue

        unit = norm_text(ws.cell(row, unit_col).value)
        next_materials = []
        next_row = row + 1
        while next_row <= ws.max_row:
            next_material = norm_text(ws.cell(next_row, material_col).value)
            next_unit = norm_text(ws.cell(next_row, unit_col).value)
            next_usage = ws.cell(next_row, usage_col).value

            if not next_material or next_unit != unit or not same_number(usage_cell.value, next_usage):
                break

            simplified = simplify_next_material(next_material)
            next_color = norm_text(ws.cell(next_row, color_col).value)
            if simplified:
                if next_color:
                    simplified = f"{simplified}*{next_color}"
                next_materials.append(simplified)
            next_row += 1

        if not next_materials:
            row += 1
            continue

        combined = "+".join(next_materials)
        old_material = norm_text(material_cell.value)
        if f"({combined})" in old_material:
            row = next_row
            continue

        new_material = f"{material}({combined})"
        if material_cell.value == new_material:
            row = next_row
            continue

        material_cell.value = new_material
        if highlight:
            material_cell.fill = HIGHLIGHT_FILL
            material_cell.comment = Comment(
                "自动优化：因本行与后续连续行单件用料相同，已把后续连续行材料名称追加到本行材料名称中。",
                "Codex",
            )

        changes.append(
            {
                "sheet": ws.title,
                "row": row,
                "old": old_material,
                "new": new_material,
                "next_rows": f"{row + 1}-{next_row - 1}",
            }
        )
        row = next_row

    return changes


def build_sheet_usage_summary(ws, include_source_rows=True, matched_groups_only=False, trigger_keywords=None):
    trigger_keywords = trigger_keywords or []
    headers = find_table_headers(ws)
    if not headers:
        return 0

    groups = {}
    matched_rows = (
        collect_matched_group_rows(ws, headers, trigger_keywords)
        if matched_groups_only
        else None
    )

    for row in range(headers["header_row"] + 1, ws.max_row + 1):
        if matched_groups_only and row not in matched_rows:
            continue

        material = norm_text(cell_value(ws, row, headers["material_col"]))
        color = norm_text(cell_value(ws, row, headers["color_col"]))
        unit = norm_text(cell_value(ws, row, headers["unit_col"]))
        spec = norm_text(cell_value(ws, row, headers["spec_col"]))
        order_usage = displayed_number(ws, row, headers["order_usage_col"])

        if order_usage is None or math.isclose(order_usage, 0):
            continue
        if not material:
            continue

        key = (material, color, unit, spec)
        if key not in groups:
            groups[key] = {"total": 0.0, "sources": []}
        groups[key]["total"] += order_usage
        groups[key]["sources"].append(f"{ws.title}!{row}")

    summary_name = f"{ws.title}-材料用量汇总"
    if summary_name in ws.parent.sheetnames:
        del ws.parent[summary_name]

    summary_ws = ws.parent.create_sheet(title=summary_name)
    summary_ws.sheet_view.showGridLines = False

    header_list = ["材料名称", "颜色", "单位", "门幅/规格", "订单用量合计"]
    if include_source_rows:
        header_list.append("来源行号")

    summary_ws.append(header_list)
    for material, color, unit, spec in sorted(groups.keys(), key=lambda item: (item[0], item[1], item[2], item[3])):
        row_data = [material, color, unit, spec, groups[(material, color, unit, spec)]["total"]]
        if include_source_rows:
            row_data.append(", ".join(groups[(material, color, unit, spec)]["sources"]))
        summary_ws.append(row_data)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")
    header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for cell in summary_ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in summary_ws.iter_rows(min_row=2, max_row=summary_ws.max_row, min_col=1, max_col=summary_ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for row in range(2, summary_ws.max_row + 1):
        summary_ws.cell(row, 5).number_format = "0.000000"

    widths = [46, 12, 10, 16, 16, 32]
    for index, width in enumerate(widths[: summary_ws.max_column], 1):
        summary_ws.column_dimensions[summary_ws.cell(1, index).column_letter].width = width
    summary_ws.freeze_panes = "A2"

    return len(groups)


def optimize_workbook(
    input_path: Path,
    output_path: Path,
    highlight: bool = True,
    trigger_keywords=None,
    create_summary: bool = True,
    include_source_rows: bool = True,
    summary_matched_groups_only: bool = True,
    selected_sheets=None,
):
    if trigger_keywords is None:
        trigger_keywords = load_trigger_keywords(None)

    selected_sheets = selected_sheets or []

    wb = load_workbook(input_path)
    changes = []

    for sheet_name in wb.sheetnames.copy():
        ws = wb[sheet_name]
        if selected_sheets and sheet_name not in selected_sheets:
            continue

        optimized_name = f"{sheet_name}-已优化"
        if optimized_name in wb.sheetnames:
            del wb[optimized_name]

        optimized_ws = copy_sheet(ws, wb, optimized_name)
        sheet_changes = optimize_sheet(optimized_ws, highlight=highlight, trigger_keywords=trigger_keywords)
        changes.extend(sheet_changes)

        if create_summary:
            build_sheet_usage_summary(
                optimized_ws,
                include_source_rows=include_source_rows,
                matched_groups_only=summary_matched_groups_only,
                trigger_keywords=trigger_keywords,
            )

    original_sheets = [name for name in wb.sheetnames if "-已优化" not in name and "-材料用量汇总" not in name]
    for sheet_name in original_sheets:
        del wb[sheet_name]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return {"changes": changes, "summary_count": len(changes)}


def choose_file_with_dialog():
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()
    path = filedialog.askopenfilename(
        title="请选择需要优化的工艺表",
        filetypes=[("Excel 文件", "*.xlsx *.xlsm"), ("所有文件", "*.*")],
    )
    if not path:
        return None
    return Path(path)


def show_done_message(output_path: Path, changes):
    try:
        import tkinter as tk
        from tkinter import messagebox

        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo(
            "工艺表格自动优化完成",
            f"已生成：\n{output_path}\n\n本次优化 {len(changes)} 处。",
        )
    except Exception:
        pass


def parse_args(argv):
    parser = argparse.ArgumentParser(description="工艺表格自动优化工具")
    parser.add_argument("input", nargs="?", help="需要优化的 Excel 文件")
    parser.add_argument("output", nargs="?", help="输出文件，可不填")
    parser.add_argument("--config", help="规则配置文件，默认使用脚本同目录 optimize_rules.json")
    parser.add_argument("--no-summary", action="store_true", help="不生成订单用量汇总工作表")
    parser.add_argument("--no-source-rows", action="store_true", help="汇总表不显示来源行号")
    parser.add_argument("--summary-all-rows", action="store_true", help="汇总全表，不限定连续同单位同用量材料组")
    parser.add_argument("--no-highlight", action="store_true", help="不标绿、不添加批注")
    return parser.parse_args(argv)


def main(argv=None):
    args = parse_args(argv or sys.argv[1:])

    if args.input:
        input_path = Path(args.input)
    else:
        input_path = choose_file_with_dialog()
        if input_path is None:
            return 0

    if not input_path.exists():
        raise FileNotFoundError(f"找不到文件：{input_path}")

    if args.output:
        output_path = Path(args.output)
    else:
        output_path = unique_output_path(input_path.with_name(f"{input_path.stem}_已优化{input_path.suffix}"))

    trigger_keywords = load_trigger_keywords(Path(args.config) if args.config else None)
    result = optimize_workbook(
        input_path=input_path,
        output_path=output_path,
        highlight=not args.no_highlight,
        trigger_keywords=trigger_keywords,
        create_summary=not args.no_summary,
        include_source_rows=not args.no_source_rows,
        summary_matched_groups_only=not args.summary_all_rows,
    )
    changes = result["changes"]

    print(f"已生成：{output_path}")
    print(f"本次优化 {len(changes)} 处。")
    if result["summary_count"] is not None:
        print(f"订单用量汇总 {result['summary_count']} 条。")
    for item in changes:
        print(
            f"- {item['sheet']} 第 {item['row']} 行："
            f"{item['old']} -> {item['new']}"
        )

    if not args.input:
        show_done_message(output_path, changes)

    return 0


NORMALIZE_HEADERS = ["产品名称", "材料名称", "颜色", "单位", "门幅/规格", "单件用料", "订单数量", "订单用量"]


def unmerge_and_fill(ws):
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        top_left_cell = ws.cell(merged_range.min_row, merged_range.min_col)
        value = top_left_cell.value
        ws.unmerge_cells(str(merged_range))
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                ws.cell(row, col).value = value


def find_header_row(ws):
    max_scan_row = min(ws.max_row, 30)
    for row in range(1, max_scan_row + 1):
        for col in range(1, ws.max_column + 1):
            value = str(ws.cell(row, col).value).strip() if ws.cell(row, col).value else ""
            if value == "材料名称":
                return row
    return None


def find_column_indices(ws, header_row):
    indices = {"产品名称": 1}
    for col in range(1, ws.max_column + 1):
        value = str(ws.cell(header_row, col).value).strip() if ws.cell(header_row, col).value else ""
        if value in NORMALIZE_HEADERS:
            indices[value] = col
    return indices


def update_formula(formula, old_row, new_row, col_mapping):
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula

    def replace_ref(match):
        ref = match.group(0)
        col_dollar = "$" if ref.startswith("$") else ""
        remaining = ref[1:] if ref.startswith("$") else ref
        col_part = re.search(r"[A-Za-z]+", remaining).group()
        after_col = remaining[len(col_part):]
        row_dollar = "$" if after_col.startswith("$") else ""
        row_part = re.search(r"\d+", after_col).group()
        old_r = int(row_part)
        new_r = old_r - (old_row - 1)

        old_col = column_index_from_string(col_part)
        if old_col in col_mapping:
            new_col = col_mapping[old_col]
            from openpyxl.utils import get_column_letter
            col_part = get_column_letter(new_col)

        return f"{col_dollar}{col_part}{row_dollar}{new_r}"

    new_formula = re.sub(r"\$?[A-Za-z]+\$?\d+", replace_ref, formula)
    return new_formula


def normalize_sheet(ws):
    unmerge_and_fill(ws)

    header_row = find_header_row(ws)
    if not header_row:
        return None

    col_indices = find_column_indices(ws, header_row)

    new_ws = ws.parent.create_sheet(title=f"{ws.title}_整理后")
    new_ws.append(NORMALIZE_HEADERS)

    col_mapping = {}
    for idx, header in enumerate(NORMALIZE_HEADERS):
        if header in col_indices:
            col_mapping[col_indices[header]] = idx + 1

    for row in range(header_row + 1, ws.max_row + 1):
        new_row = []
        for idx, header in enumerate(NORMALIZE_HEADERS):
            if header in col_indices:
                old_col = col_indices[header]
                cell = ws.cell(row, old_col)
                value = cell.value
                if header == "订单用量" and isinstance(value, str) and value.startswith("="):
                    value = update_formula(value, header_row, 1, col_mapping)
                new_row.append(value)
            else:
                new_row.append(None)
        new_ws.append(new_row)

    new_ws.title = ws.title
    return new_ws


def normalize_workbook(input_path: Path, output_path: Path):
    wb = load_workbook(input_path)
    sheets_to_process = wb.sheetnames.copy()

    for sheet_name in sheets_to_process:
        ws = wb[sheet_name]
        new_ws = normalize_sheet(ws)
        if new_ws:
            del wb[sheet_name]
            new_ws.title = sheet_name

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


if __name__ == "__main__":
    raise SystemExit(main())
