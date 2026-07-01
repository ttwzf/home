# -*- coding: utf-8 -*-

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from optimize_process_table import add_conditional_format, find_headers


def build_template(src: Path, out: Path):
    wb = load_workbook(src)

    for ws in wb.worksheets:
        headers = find_headers(ws)
        if headers:
            add_conditional_format(ws, headers[0], headers[1])

    if "工具说明" in wb.sheetnames:
        del wb["工具说明"]

    ws = wb.create_sheet("工具说明", 0)
    ws.sheet_view.showGridLines = False

    ws["A1"] = "工艺表格自动优化模板"
    ws["A1"].font = Font(name="Microsoft YaHei", size=18, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="305496")
    ws.merge_cells("A1:F1")

    rows = [
        (
            "处理规则",
            "第一行材料名称包含配置文件中的触发关键词时，向下查找连续且单件用料、单位都相同的行，保持第一行材料名称不变，把后续连续行材料名称去掉“（复合）”后，按“材料名称*颜色”依次追加到第一行括号内。",
        ),
        ("保留内容", "不删除任何行；订单用量和原公式保留，不重新计算。"),
        (
            "输出方式",
            "双击同目录下“运行工艺表格自动优化工具.bat”，选择要处理的 Excel 文件，会另存为“原文件名_已优化.xlsx”。",
        ),
        (
            "关键词配置",
            "同目录下 optimize_rules.json 里的 trigger_keywords 可填写多个触发关键词；默认是“需复合”。如果改成空列表 []，任意非空材料名称都可作为第一行。",
        ),
        (
            "汇总分支",
            "可视化工具默认勾选“只汇总连续同用量材料组”，汇总表只统计参与连续组的材料行；取消勾选后按全表有效材料行汇总。",
        ),
        ("按钮说明", "纯 .xlsx 不能保存可执行宏按钮；如需 Excel 内按钮，需要另做 .xlsm 宏版。"),
        ("已优化数量", '=COUNTIF(\'盘套 (2)\'!B5:B140,"*需复合(*")'),
    ]

    for row_index, (label, value) in enumerate(rows, 3):
        ws.cell(row_index, 1).value = label
        ws.cell(row_index, 2).value = value

    ws["A10"] = "识别提示"
    ws["B10"] = "数据表中材料名称包含“需复合(”的单元格会以浅绿色条件格式提示。"

    thin = Side(style="thin", color="D9E2F3")
    for row in ws.iter_rows(min_row=3, max_row=10, min_col=1, max_col=2):
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.fill = PatternFill("solid", fgColor="F8FBFF")
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in range(3, 11):
        ws.cell(row, 1).font = Font(name="Microsoft YaHei", bold=True, color="1F4E79")
        ws.cell(row, 2).font = Font(name="Microsoft YaHei")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 92
    for row in range(3, 11):
        ws.row_dimensions[row].height = 38

    wb.save(out)


if __name__ == "__main__":
    build_template(Path("盘套工艺表.xlsx"), Path("工艺表格自动优化模板.xlsx"))
