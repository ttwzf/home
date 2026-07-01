# -*- coding: utf-8 -*-
"""
工艺表格自动优化 - 可视化工具
"""

from __future__ import annotations

import json
import threading
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

from optimize_process_table import (
    DEFAULT_TRIGGER_KEYWORDS,
    app_base_dir,
    load_trigger_keywords,
    normalize_workbook,
    optimize_workbook,
    unique_output_path,
)


APP_DIR = app_base_dir()
CONFIG_PATH = APP_DIR / "optimize_rules.json"


class OptimizerApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("工艺表格自动优化工具")
        self.geometry("800x550")
        self.minsize(700, 500)

        self.input_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.status_text = tk.StringVar(value="请选择需要优化的 Excel 文件。")
        self.highlight = tk.BooleanVar(value=True)
        self.create_summary = tk.BooleanVar(value=True)
        self.include_source_rows = tk.BooleanVar(value=True)
        self.summary_matched_groups_only = tk.BooleanVar(value=False)
        self.selected_sheets = []

        self._build_ui()
        self._load_keywords()

    def _build_ui(self):
        self.columnconfigure(0, weight=1)
        self.rowconfigure(4, weight=1)

        title = ttk.Label(self, text="工艺表格自动优化工具", font=("Microsoft YaHei", 16, "bold"))
        title.grid(row=0, column=0, sticky="w", padx=15, pady=(12, 8))

        file_frame = ttk.LabelFrame(self, text="文件")
        file_frame.grid(row=1, column=0, sticky="ew", padx=15, pady=4)
        file_frame.columnconfigure(1, weight=1)

        ttk.Label(file_frame, text="输入文件").grid(row=0, column=0, sticky="w", padx=8, pady=4)
        ttk.Entry(file_frame, textvariable=self.input_path).grid(row=0, column=1, sticky="ew", padx=8, pady=4)
        ttk.Button(file_frame, text="选择", command=self.choose_input, width=6).grid(row=0, column=2, padx=8, pady=4)

        ttk.Label(file_frame, text="输出文件").grid(row=1, column=0, sticky="w", padx=8, pady=4)
        ttk.Entry(file_frame, textvariable=self.output_path).grid(row=1, column=1, sticky="ew", padx=8, pady=4)
        ttk.Button(file_frame, text="另存为", command=self.choose_output, width=6).grid(row=1, column=2, padx=8, pady=4)

        mid_frame = ttk.Frame(self)
        mid_frame.grid(row=2, column=0, sticky="ew", padx=15, pady=4)
        mid_frame.columnconfigure(0, weight=1)
        mid_frame.columnconfigure(1, weight=1)

        sheet_frame = ttk.LabelFrame(mid_frame, text="工作表选择")
        sheet_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 6))
        sheet_frame.columnconfigure(0, weight=1)

        ttk.Label(sheet_frame, text="Ctrl+点击可多选").grid(row=0, column=0, sticky="w", padx=8, pady=(4, 2))
        self.sheet_listbox = tk.Listbox(sheet_frame, selectmode=tk.EXTENDED, height=4)
        self.sheet_listbox.grid(row=1, column=0, sticky="ew", padx=8, pady=2)
        sheet_btn_frame = ttk.Frame(sheet_frame)
        sheet_btn_frame.grid(row=2, column=0, sticky="w", padx=8, pady=(2, 4))
        ttk.Button(sheet_btn_frame, text="全选", command=self.select_all_sheets, width=5).pack(side="left")
        ttk.Button(sheet_btn_frame, text="取消", command=self.deselect_all_sheets, width=5).pack(side="left", padx=5)

        rule_frame = ttk.LabelFrame(mid_frame, text="触发关键词")
        rule_frame.grid(row=0, column=1, sticky="nsew", padx=(6, 0))
        rule_frame.columnconfigure(0, weight=1)

        help_text = "包含关键词才触发；留空表示任意材料"
        ttk.Label(rule_frame, text=help_text).grid(row=0, column=0, sticky="w", padx=8, pady=(4, 2))

        self.keyword_text = tk.Text(rule_frame, height=3, font=("Microsoft YaHei", 10))
        self.keyword_text.grid(row=1, column=0, sticky="ew", padx=8, pady=2)

        button_row = ttk.Frame(rule_frame)
        button_row.grid(row=2, column=0, sticky="ew", padx=8, pady=(2, 4))
        ttk.Button(button_row, text="保存", command=self.save_keywords, width=5).pack(side="left")
        ttk.Button(button_row, text="默认", command=self.reset_keywords, width=5).pack(side="left", padx=5)

        option_frame = ttk.Frame(self)
        option_frame.grid(row=3, column=0, sticky="ew", padx=15, pady=4)

        ttk.Checkbutton(option_frame, text="标绿并添加批注", variable=self.highlight).pack(side="left", padx=(0, 12))
        ttk.Checkbutton(option_frame, text="生成订单用量汇总表", variable=self.create_summary).pack(side="left", padx=(0, 12))
        ttk.Checkbutton(option_frame, text="只汇总连续同用量材料组", variable=self.summary_matched_groups_only).pack(side="left", padx=(0, 12))
        ttk.Checkbutton(option_frame, text="汇总表显示来源行号", variable=self.include_source_rows).pack(side="left", padx=(0, 12))

        log_frame = ttk.LabelFrame(self, text="处理结果")
        log_frame.grid(row=4, column=0, sticky="nsew", padx=15, pady=4)
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)

        self.log = tk.Text(log_frame, font=("Consolas", 10), wrap="word")
        self.log.grid(row=0, column=0, sticky="nsew")
        scrollbar = ttk.Scrollbar(log_frame, orient="vertical", command=self.log.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.log.configure(yscrollcommand=scrollbar.set)

        action_frame = ttk.Frame(self)
        action_frame.grid(row=4, column=0, sticky="ew", padx=15, pady=(4, 12))
        ttk.Button(action_frame, text="整理表格", command=self.run_normalize).pack(side="left")
        ttk.Button(action_frame, text="开始优化", command=self.run_optimize).pack(side="left", padx=8)
        ttk.Button(action_frame, text="打开输出文件夹", command=self.open_output_folder).pack(side="left", padx=8)
        status_label = ttk.Label(action_frame, textvariable=self.status_text, foreground="#666666")
        status_label.pack(side="left", padx=12)

    def _load_keywords(self):
        try:
            keywords = load_trigger_keywords(CONFIG_PATH)
        except Exception:
            keywords = DEFAULT_TRIGGER_KEYWORDS
        self.keyword_text.delete("1.0", "end")
        self.keyword_text.insert("1.0", "\n".join(keywords))

    def _get_keywords_from_ui(self):
        raw = self.keyword_text.get("1.0", "end").splitlines()
        return [line.strip() for line in raw if line.strip()]

    def _write_log(self, text):
        self.log.insert("end", text + "\n")
        self.log.see("end")

    def _load_sheet_list(self):
        input_path = Path(self.input_path.get().strip())
        if not input_path.exists():
            return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(input_path)
            self.sheet_listbox.delete(0, tk.END)
            for sheet_name in wb.sheetnames:
                self.sheet_listbox.insert(tk.END, sheet_name)
            self.sheet_listbox.selection_set(0, tk.END)
        except Exception as exc:
            self._write_log(f"读取工作表列表失败：{exc}")

    def select_all_sheets(self):
        self.sheet_listbox.selection_set(0, tk.END)

    def deselect_all_sheets(self):
        self.sheet_listbox.selection_clear(0, tk.END)

    def get_selected_sheets(self):
        selected_indices = self.sheet_listbox.curselection()
        return [self.sheet_listbox.get(idx) for idx in selected_indices]

    def choose_input(self):
        path = filedialog.askopenfilename(
            title="请选择需要优化的 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm"), ("所有文件", "*.*")],
        )
        if not path:
            return
        input_path = Path(path)
        self.input_path.set(str(input_path))
        self.output_path.set(str(unique_output_path(input_path.with_name(f"{input_path.stem}_已优化{input_path.suffix}"))))
        self._load_sheet_list()
        self.status_text.set("已选择输入文件。")

    def choose_output(self):
        initial = self.output_path.get()
        path = filedialog.asksaveasfilename(
            title="请选择输出文件",
            initialfile=Path(initial).name if initial else "工艺表_已优化.xlsx",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
        )
        if path:
            self.output_path.set(path)

    def save_keywords(self):
        keywords = self._get_keywords_from_ui()
        CONFIG_PATH.write_text(
            json.dumps(
                {
                    "trigger_keywords": keywords,
                    "note": "trigger_keywords 为空列表 [] 时，任意非空材料名称都可作为第一行。",
                },
                ensure_ascii=False,
                indent=2,
            ),
            encoding="utf-8",
        )
        self.status_text.set("关键词已保存。")
        self._write_log(f"已保存触发关键词：{keywords if keywords else '任意材料'}")

    def reset_keywords(self):
        self.keyword_text.delete("1.0", "end")
        self.keyword_text.insert("1.0", "\n".join(DEFAULT_TRIGGER_KEYWORDS))
        self.save_keywords()

    def run_normalize(self):
        input_path = Path(self.input_path.get().strip())
        output_path = Path(self.output_path.get().strip())
        if not input_path.exists():
            messagebox.showwarning("缺少输入文件", "请先选择需要优化的 Excel 文件。")
            return
        if not output_path:
            messagebox.showwarning("缺少输出文件", "请先选择输出文件位置。")
            return

        output_path = unique_output_path(output_path)
        self.output_path.set(str(output_path))
        self.status_text.set("正在整理表格...")
        self._write_log("")
        self._write_log(f"开始整理表格：{input_path}")

        thread = threading.Thread(
            target=self._run_normalize_worker,
            args=(input_path, output_path),
            daemon=True,
        )
        thread.start()

    def _run_normalize_worker(self, input_path, output_path):
        try:
            normalize_workbook(input_path=input_path, output_path=output_path)
        except Exception as exc:
            self.after(0, self._on_error, exc)
            return
        self.after(0, self._on_normalize_done, output_path)

    def _on_normalize_done(self, output_path):
        self.status_text.set("表格整理完成。")
        self._write_log(f"已生成：{output_path}")
        self._load_sheet_list()
        messagebox.showinfo("完成", f"表格整理完成：\n{output_path}")

    def run_optimize(self):
        input_path = Path(self.input_path.get().strip())
        output_path = Path(self.output_path.get().strip())
        if not input_path.exists():
            messagebox.showwarning("缺少输入文件", "请先选择需要优化的 Excel 文件。")
            return
        if not output_path:
            messagebox.showwarning("缺少输出文件", "请先选择输出文件位置。")
            return

        selected_sheets = self.get_selected_sheets()
        if not selected_sheets:
            messagebox.showwarning("未选择工作表", "请至少选择一个工作表进行优化。")
            return

        output_path = unique_output_path(output_path)
        self.output_path.set(str(output_path))
        self.save_keywords()
        keywords = self._get_keywords_from_ui()
        self.status_text.set("正在处理...")
        self._write_log("")
        self._write_log(f"开始处理：{input_path}")
        self._write_log(f"选择的工作表：{selected_sheets}")

        thread = threading.Thread(
            target=self._run_optimize_worker,
            args=(input_path, output_path, keywords, selected_sheets),
            daemon=True,
        )
        thread.start()

    def _run_optimize_worker(self, input_path, output_path, keywords, selected_sheets):
        try:
            normalize_workbook(input_path=input_path, output_path=output_path)
            self.after(0, self._write_log, "表格整理完成，开始优化...")
            changes = optimize_workbook(
                input_path=output_path,
                output_path=output_path,
                highlight=self.highlight.get(),
                trigger_keywords=keywords,
                create_summary=self.create_summary.get(),
                include_source_rows=self.include_source_rows.get(),
                summary_matched_groups_only=self.summary_matched_groups_only.get(),
                selected_sheets=selected_sheets,
            )
        except Exception as exc:
            self.after(0, self._on_error, exc)
            return
        self.after(0, self._on_optimize_done, output_path, changes)

    def _on_optimize_done(self, output_path, result):
        changes = result["changes"]
        summary_count = result["summary_count"]
        self.status_text.set(f"完成：优化 {len(changes)} 处。")
        self._write_log(f"已生成：{output_path}")
        self._write_log(f"本次优化 {len(changes)} 处。")
        if summary_count is not None:
            self._write_log(f"订单用量汇总 {summary_count} 条。")
        for item in changes:
            self._write_log(f"- {item['sheet']} 第 {item['row']} 行：{item['old']} -> {item['new']}")
        extra = f"\n订单用量汇总 {summary_count} 条。" if summary_count is not None else ""
        messagebox.showinfo("完成", f"已生成：\n{output_path}\n\n本次优化 {len(changes)} 处。{extra}")

    def _on_error(self, exc):
        self.status_text.set("处理失败。")
        self._write_log(f"处理失败：{exc}")
        messagebox.showerror("处理失败", str(exc))

    def open_output_folder(self):
        path_text = self.output_path.get().strip() or self.input_path.get().strip()
        if not path_text:
            return
        folder = Path(path_text).parent
        if folder.exists():
            import os
            os.startfile(folder)


if __name__ == "__main__":
    app = OptimizerApp()
    app.mainloop()
